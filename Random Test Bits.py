
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
ExcelFileName = 'AutomationTest.xlsx'
wb = load_workbook(ExcelFileName)
ws = wb.active
ws['L4'].fill = PatternFill(fgColor='FF4229', fill_type='solid')
wb.save(ExcelFileName)
