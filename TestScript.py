from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC, wait
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException, NoSuchElementException
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
import tkinter.font as tkFont
from tkinter import *
from tkinter import messagebox, LabelFrame, Frame, filedialog, ttk
from ttkbootstrap import Style, Colors
import os


def LoadingExcelInfo():
    global sheetindex, questionType, QuestionName, InstructionsToBeEntered, HelpTextToBeEntered, DropTargetToBeEntered, CorrectAnswerCell, Objective, SubObjective, TotalSheets
    global correctanswerAmount, amountofAnswers, answerIndex, allAnswersClicked, correctAnswerIndex, DragOptionToBeEntered, ws, i, wb, sheets, CorrectAnswerCellList, ObjectivesCell
    #ExcelFileName = ''
    wb = load_workbook(ExcelFileName)
    sheetindex = 0
    sheets = wb.sheetnames
    print("SheetNames: " + str(sheets))
    ws = wb.active
    i = 4
    questionType = ws['C'+str(i)]
    QuestionName = ws['B'+str(i)]
    InstructionsToBeEntered = ws['D'+str(i)]
    HelpTextToBeEntered = ws['H'+str(i)]
    DragOptionToBeEntered = ws['E'+str(i)]
    DropTargetToBeEntered = 'Drop Target Entered'
    CorrectAnswerCell = ws['F'+str(i)]
    CorrectAnswerCellList = ws['F'+str(i)]
    ObjectivesCell = ws['J'+str(i)]
    ObjectivesCell = str(ObjectivesCell.value).split('\n')
    Objective = str(ObjectivesCell[0])
    SubObjective = str(ObjectivesCell[1])
    TotalSheets = len(sheets)
    correctanswerAmount = len(str(CorrectAnswerCellList.value).split('\n'))

    amountofAnswers = len(str(DragOptionToBeEntered.value).split('\n'))
    answerIndex = 0
    allAnswersClicked = 0
    correctAnswerIndex = 0
    #print("Amount of Total Answers"+str(amountofAnswers))


def SheetChecker():
    global i, questionType, QuestionName, InstructionsToBeEntered, HelpTextToBeEntered, DropTargetToBeEntered, CorrectAnswerCell, Objective, SubObjective, correctanswerAmount, amountofAnswers
    print(str(i)+" before")
    i += 1
    if ws.max_row != (i-1):
        print(str(ws.max_row) + ' Max Row Number')
        global CorrectAnswerCellList, ObjectivesCell
        questionType = ws['C'+str(i)]
        QuestionName = ws['B'+str(i)]
        InstructionsToBeEntered = ws['D'+str(i)]
        HelpTextToBeEntered = ws['H'+str(i)]
        DragOptionToBeEntered = ws['E'+str(i)]
        DropTargetToBeEntered = 'Drop Target Entered'
        CorrectAnswerCell = ws['F'+str(i)]
        CorrectAnswerCellList = ws['F'+str(i)]
        ObjectivesCell = ws['J'+str(i)]
        ObjectivesCell = str(ObjectivesCell.value).split('\n')
        Objective = str(ObjectivesCell[0])
        SubObjective = str(ObjectivesCell[1])
        correctanswerAmount = len(str(CorrectAnswerCellList.value).split('\n'))
        amountofAnswers = len(str(DragOptionToBeEntered.value).split('\n'))
        print(str(i)+" after")
        QuestionTypeClicker()
        QuestionCreation()
    else:
        print("Finished cycling!!")
    # print(str(DragOptionToBeEntered.value).split('\n'))
    #print(str(i) + " :This is the i value")


def TimeoutErrorMessage():
    global QIDCell, i
    QIDCell = ws['N'+str(i)]
    QIDCell.fill = PatternFill(fgColor='34B1EB', fill_type='solid')
    QIDCell.value = "Question Failed To Create:" + ErrorMessage
    wb.save(ExcelFileName)


def ClickMoreAnswer():
    global driver
    if amountofAnswers > answerIndex:
        # Create Another Answer
        CreateAnotherAnswer = driver.find_element_by_id(
            "AddAnswer0"+str(answerIndex))
        CreateAnotherAnswer.click()


def FinallyCreateQuestionButton():
    global driver, questionType
    if questionType.value == 'IFrame':
        FinalCreate = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="createQuestion"]/div[3]/div/input'))
        ).click()
    else:
        FinalCreate = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, 'OnClickCreateQuestion'))
        ).click()
    time.sleep(2)


def RetreiveQID():
    global driver, i, QIDCell, QIDName
    QIDName = driver.find_element_by_class_name("modal-body").text
    QIDCell = ws['N'+str(i)]
    if questionType.value == "Multiple Choice":
        QIDCell.fill = PatternFill(fgColor='29FF49', fill_type='solid')
    else:
        QIDCell.fill = PatternFill(fgColor='FF4229', fill_type='solid')

    QIDCell.value = str(QIDName).strip(
        "Question successfully created. Question ID: ")
    wb.save(ExcelFileName)
    print(QIDName)


def QuestionTypeClicker():
    global driver, questionType
    openQuestionDD = WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.XPATH, ('//*[@id="questionTypeContainer"]/div[2]/span[1]/span/span[1]'))))
    openQuestionDD.click()
    if questionType.value.lower() == 'multiple choice' or questionType.value.lower() == 'choose' or questionType.value.lower() == 'multiple' or questionType.value.lower() == 'mcwi':
        questionType.value = 'Multiple Choice'
    elif questionType.value.lower() == "dnd" or questionType.value.lower() == "drag and match":
        questionType.value = 'Drag and Match'
    elif questionType.value.lower() == 'applications':
        questionType.value = 'Application'
    elif questionType.value.lower() == 'demo' or questionType.value.lower() == 'iframe' or questionType.value == 'Demo':
        questionType.value = 'IFrame'
    time.sleep(1)
    questionTypeSelected = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="QuestionType-list"]//li[text() = "'+questionType.value+'"]'
                                    ))
    )
    print("Question Type: " + questionType.value)
    questionTypeSelected.click()


def QuestionCreation():
    global questionType, QuestionName, answerIndex, amountofAnswers, DragOptionToBeEntered, DropTargetToBeEntered, InstructionsToBeEntered, HelpTextToBeEntered, i, driver, allAnswersClicked, Objective, SubObjective, correctanswerAmount
    global correctAnswerIndex, CorrectAnswerCell, QIDCell, ErrorMessage, i
    # Inputting Question Name
    time.sleep(1)
    questionNameEntry = WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.ID, 'QuestionName')))
    #questionNameEntry = driver.find_element_by_id("QuestionName")
    questionNameEntry.send_keys(QuestionName.value)
    # Opens Objective Menu
    objectiveDD = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="ObjectiveId"]'
                                        ))
    )
    objectiveDD.click()

    time.sleep(1)
    try:
        selectionOBJ = Select(driver.find_element_by_xpath(
            "//select[@name='ObjectiveId']"))
        selectionOBJ.select_by_visible_text(Objective.strip(" "))
    except:
        print("Object failed to Match any listed under this product")
        ErrorMessage = "Object failed to Match any listed under this product"
        TimeoutErrorMessage()
        ErrorWindowDefault()
        SheetChecker()
        driver.quit()
        ResetWindow()
        LoginAndOpenQuestionInput()
    # Opens subObjective Drop Down
    subObjectiveDD = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located(
            (By.XPATH, '//*[@id="ddlSubObjective"]'))
    )
    subObjectiveDD.click()
    time.sleep(1)
    selectionSubOBJ = Select(driver.find_element_by_xpath(
        "//select[@name='ddlSubObjective']"))
    try:
        selectionSubOBJ.select_by_visible_text(SubObjective.strip(" "))
    except TimeoutException or ElementClickInterceptedException or NoSuchElementException:
        ErrorMessage = "SubObjective does not match any listed under this product"
        ErrorWindowDefault()
        TimeoutErrorMessage()
        SheetChecker()
        driver.quit()
        ResetWindow()
        print("SubObjective does not match any listed under this product")

    time.sleep(1)
    # Instructions
    InstructionFrame = WebDriverWait(driver, 20).until(
        EC.frame_to_be_available_and_switch_to_it((By.XPATH, '//*[@id="cke_1_contents"]/iframe')))
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
        (By.XPATH, "//body[@class='cke_editable cke_editable_themed cke_contents_ltr cke_show_borders']/p"))).send_keys(InstructionsToBeEntered.value)
    driver.switch_to.default_content()
    # HelpText
    HelpFrame = WebDriverWait(driver, 20).until(
        EC.frame_to_be_available_and_switch_to_it((By.XPATH, '//*[@id="cke_2_contents"]/iframe')))
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
        (By.XPATH, "//body[@class='cke_editable cke_editable_themed cke_contents_ltr cke_show_borders']/p"))).send_keys(HelpTextToBeEntered.value)
    driver.switch_to.default_content()

    if questionType.value == 'Multiple Choice':
        DragOptionToBeEntered = str(DragOptionToBeEntered.value).split('\n')
        CorrectAnswerCell = str(CorrectAnswerCell.value).split('\n')
        try:
            while amountofAnswers > answerIndex:
                MultipleChoice1 = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((
                        By.XPATH, "//input[@name='InstructionComponent[0].AnswersBlock.Answers["+str(answerIndex)+"].AnswerText']"))).send_keys(DragOptionToBeEntered[answerIndex])
                if correctanswerAmount != allAnswersClicked:
                    if str(CorrectAnswerCell[correctAnswerIndex]) == str(DragOptionToBeEntered[answerIndex]):
                        CorrectAnswerClicked = WebDriverWait(driver, 20).until(
                            EC.element_to_be_clickable(
                                (By.XPATH, '//*[@id="InstructionComponent_0__AnswersBlock_Answers_'+str(answerIndex)+'__IsCorrect"]'))
                        ).click()
                        correctAnswerIndex += 1
                        allAnswersClicked += 1
                        time.sleep(1)
                if (int(amountofAnswers)-1) != answerIndex:
                    ClickMoreAnswer()
                    time.sleep(1)
                    #print(str(answerIndex) + "current answer index")
                answerIndex += 1
        except:
            ErrorMessage = "Answers Formatted Incorrectly"
            ErrorWindowDefault()
            TimeoutErrorMessage()
            SheetChecker()
            driver.quit()
            ResetWindow()
            print("Answers Formatted Incorrectly")

    FinallyCreateQuestionButton()
    RetreiveQID()
    CloseQIDMenu = WebDriverWait(driver, 15).until(
        EC.presence_of_element_located(
            (By.XPATH, '//*[@id="successfulUpdateModal"]/div/div/div[1]/button/span'))
    )
    CloseQIDMenu.click()
    SheetChecker()


def LoginAndOpenQuestionInput():
    global Username, Password, driver, categoryName, productName, ErrorMessage, ErrorWindow
    if not driver:

        Username = Username_var.get()
        Password = Password_var.get()
        categoryName = " " + Category_var.get()
        productName = Product_var.get()
        PATH = "C:\Program Files (x86)\chromedriver.exe"
        driver = webdriver.Chrome(PATH)
        driver.get("https://author.gmetrix.net")
        # This is opening the webpage^^^
        driver.maximize_window()
        search = driver.find_element_by_id("Email")
        # Your Credentials here
        search.send_keys(Username)
        search = driver.find_element_by_id("Password")
        search.send_keys(Password)
        time.sleep(4)

        signIn = driver.find_element_by_xpath('//*[@id="buttonSignIn"]')
        signIn.click()
        # This is Signing the user in ^^

        questionTab = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.LINK_TEXT, "Questions"))
        )
        questionTab.click()
        createButton = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.LINK_TEXT, "Create"))
        )
        createButton.click()
        # Opening Category Menu
        openCategoryDD = driver.find_element_by_css_selector(
            'span[class="k-input"]')
        openCategoryDD.click()
        try:
            # Clicking Category Selected
            categorySelected = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="Category-list"]//li[text() = "'+categoryName+'"]'
                                            ))
            )
            categorySelected.click()
        except:
            ErrorMessage = "Category Entry Failed"
            ErrorWindowDefault()
            driver.quit()
            ResetWindow()
            print("Category Entry Failed")
        # Opening Product Drop Down List
        openProductDD = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, ('/html/body/div[2]/form/div[1]/div/div/div[2]/div[2]/span[1]/span/span[1]'))))
        openProductDD.click()
        time.sleep(2)
        try:
            # Clicking Product Name
            productSelected = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="Product-list"]//li[text() = "'+productName+'"]'
                                            ))
            )
            productSelected.click()
        except:
            ErrorMessage = "Product Entered Failed"
            ErrorWindowDefault()
            driver.quit()
            ResetWindow()
            print("Product Entered Failed")

        time.sleep(2)
        # Opening Question Drop Down List
        # Inputing the Question Type
        QuestionTypeClicker()
        firstCreateButton = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[2]/form/div[2]/input'))
        )
        firstCreateButton.click()
        QuestionCreation()


def UpdateProductList(event):
    global ProductInput
    selected = event.widget.get()
    ProductList_values = {
        "Adobe 2021": ["After Effects 2019/2020/2021", "Animate 2019/2020/2021", "Dreamweaver 2019/2020/2021", "Illustrator 2019/2020/2021", "InDesign 2019/2020/2021", "Photoshop 2019/2020/2021", "Premiere Pro 2019/2020/2021"],
        "Adobe CC": ["After Effects CC", "Animate CC", "Dreamweaver CC", "Flash CC", "Illustrator", "InDesign", "Photoshop", "Premiere"],
        "Adobe CC 2018 (LITA)": ["After Effects CC 2018/2019/2020", "Animate CC 2018/2019/2020", "Dreamweaver CC 2018/2019/2020", "Illustrator CC 2018/2019/2020", "InDesign CC 2018/2019/2020", 'Photoshop CC 2018/2019/2020', "Premiere Pro CC 2018/2019/2020"],
        'Adobe CS5': ["Dreamweaver CS5", "Flash CS5", "Photoshop CS5"],
        "Adobe CS6": ["Dreamweaver CS6", "Flash CS6", "Illustrator CS6", "InDesign CS6", "Photoshop CS6", "Premiere CS6"],
        "AHIT": ["AHIT Live Class", "Home Inspection Concepts", "National Home Inspector Examination"],
        "AppInventor": ["AppInventor"],
        "Apple": ["App Development with Swift - Level 1", "App Development with Swift Associate", "App Development with Swift Certified User"],
        "Appraisal": ["NAJA"],
        "ASE Entry-Level Certification Program": ["Automatic Transmission and Transaxle", "Brakes", "Electrical/Electronic Systems", "Engine Performance", "Engine Repair", "Heating, Ventilation and Air Conditioning", "Manual Drivetrains and Axels", "Suspension and Steering"],
        "ASE Professional Certification": ["A1: Engine Repair", "A2: Automatic Transmission/Transaxle", "A3: Manual Drivetrains and Axels", "A4: Suspension and Steering", "A5: Brakes", "A6: Electrical and Electronic Systems", "A7: Heating and Air Conditioning", "A8: Engine Performance", "A9: Light Vehicle Diesel Engines", "C1: Automobile Service Consultant", "F1: Compressed Natural Gas Vehicle", "G1: Auto Maintenance and Light Repair", "L1: Advanced Engine Performance Specialist", "L3: Light Duty Hybrid/Electric Vehicle Specialist", "P2: Automobile Parts Specialist"],
        "Autodesk Certified Professional": ["3DS Max", "AutoCAD", "AutoCAD Civil 3D", "AutoCAD for Design and Drafting", "Civil 3D for Infrastructure Design", "Inventor", "Inventor for Mechanical Design", "Maya", "Revit Architecture", "Revit for Architectural Design", "Revit for Structural Design", "Revit MEP", "Revit Structure"],
        "Autodesk Certified User": ["3DS Max", "AutoCAD", "Fusion 360", "Inventor", "Maya", "Revit Architecture"],
        "AWS Certified": ["Cloud Practitioner"],
        "Black Knight": ["Paragon"],
        "Coding in Minecraft": ["Coding in Minecraft", "Introduction to Coding using MakeCode"],
        "Communication Skills for Business (CSB)": ["Communication Skills for Business (CSB)", "English for IT"],
        "CompTIA": ["A+ (220-1001)", "A+ (220-1002)", "A+ 901", "A+ 902", "Advanced Security Practitioner (CAS-002)", "Advanced Security Practitioner (CAS-003)", "Cloud + (CV0-002)", "IT Fundamentals (FC0-U61)", "Linux+ LX0-103", "Linux+ LX0-104", "Network+", "Network+ (N10-007)", "Network+ (N10-008)", "PenTest + (PT0-001)", "Security+", "Security+ SY0-501", "Security+ SY0-601", "Server + (SK0-004)"],
        "Construction": ["Licensing"],
        "Digital Skills Programme": ["Digital Skills Programme"],
        "EC-Council": ["Cyber Forensics Associate", "Ethical Hacking Associate"],
        "Entrepreneurship & Small Business": ["Entrepreneurship & Small Business", "Entrepreneurship & Small Business v.2 - U.S."],
        "Exam Prep": ["Artisan Content Import", "California Real Estate Assessment", "Exam Prep Plus - National Real Estate Assessment", "Georgia Real Estate Assessment", "National AMP-PSI Real Estate Assessment", "National Pearson Vue Real Estate Assessment", "New York Real Estate Assessment", "Oregon Real Estate Assessment", "Pearson Vue + Utah Practice Exam", "Texas Real Estate Assessment"],
        "EXIN": ["Ethical Hacking"],
        "GISP": ["GISCI"],
        "Global Digital Literacy": ["Bulgarian Digital Skills", "Digital Literacy Baseline Assessment"],
        "GMetrix Competency": ["GMetrix Competency"],
        "IC PHP Developer Fundamentals": ["IC PHP Developer Fundamentals"],
        "IC3 GS4": ["Computing Fundamentals", "IC3 GS4 All Products", "IC3 GS4 Computing Fundamentals Spark", "IC3 GS4 Key Applications Spark", "IC3 GS4 Living Online Spark", "IC3 GS4 Spark", "Key Applications", "Living Online"],
        "IC3 GS5": ["Computing Fundamentals", "Computing Fundamentals (Office 2016)", "IC3 Fast Track", "IC3 GS5 Spark", "Key Applications", "Living Online"],
        "IC3 GS6": ["Digital Literacy Level 1", "Digital Literacy Level 2", "Digital Literacy Level 3"],
        "ICC Certifications": ["Property Maintenance and Housing Inspector"],
        "In Development": ["AIJR - Situational Judgement", "Am I Job Ready?", "App Inventor", "GMetrix Internal Content", "Java 9", "Learning HTML & CSS", "LearnKey Soft Skills", "NYC Taxi", "Premium Accident and Health Producer", "Premium Life, Accident, and Health Insurance Producer", "QA", "TPG Demo"],
        "Information Technology Specialist": ["Artificial Intelligence", "Cloud Computing", "Computational Thinking", "Cybersecurity", "Databases", "Device Configuration and Management", "HTML 5 Application Development", "HTML and CSS", "Java", "JavaScript", "Network Security", "Networking", "Python", "Software Development"],
        "Introduction to Programming": ["Introduction to Programming"],
        "Intuit": ["Certified Bookkeeping Professional", "Design for Delight Innovator", "QuickBooks Desktop", "QuickBooks Online", "QuickBooks Online - U.S.", "QuickBooks Online Global"],
        "Microsoft Certified Educator (MCE)": ["Microsoft Certified Educator (2018)", "Microsoft Certified Educator (Old Version)"],
        "Microsoft Certified Fundamentals": ["AI-900 AI Fundamentals", "AZ-204: Developing Solutions for Microsoft Azure", "AZ-900 Azure Fundamentals", "Azure Administrator (AZ-104)", "Azure Security Engineer (AZ-500)", "DA-100: Analyzing Data with Microsoft Power BI", "DP-900 Azure Data Fundamentals", "MB-900 Dynamics 365 Fundamentals", "MB-901 Dynamics 365 Fundamentals", "MB-910: Microsoft Dynamics 365 Fundamentals (CRM)", "MB-920: Microsoft Dynamics 365 Fundamentals (ERP)", "MS-500: Microsoft 365 Security Administration", "MS-600: Building Applications and Solutions with Microsoft 365 Core Services", "MS-700: Managing Microsoft Teams", "MS-900 Microsoft 365 Fundamentals", "PL-100: Microsoft Power Platform App Maker", "PL-900 Power Platform Fundamentals", "SC-900: Microsoft Security, Compliance, and Identity Fundamentals"],
        'Microsoft Office 2010': ["Access", "Excel", "Excel Expert", "Office 365", "OneNote", "Outlook", "PowerPoint", "SharePoint", "Word", "Word Expert"],
        'Microsoft Office 2013': ["Access", "Excel", "Excel Expert", "OneNote", "Outlook", "PowerPoint", "SharePoint", "Word", "Word Expert"],
        'Microsoft Office 2016': ["Access", "Excel", "Excel Expert", "OneNote", "Outlook", "PowerPoint", "SharePoint", "Word", "Word Expert"],
        'Microsoft Office 2019': ["Access", "Excel", "Excel Expert", "Outlook", "PowerPoint", "Word", "Word Expert"],
        "Microsoft Office Corporate Competency": ["Excel", "Teams"],
        "MTA": ["98-349: Windows Operating System Fundamentals", "98-361: Software Development Fundamentals", "98-362: Windows Development Fundamentals", "98-363: Web Development Fundamentals", "98-364: Database Administration Fundamentals", "98-365: Windows Server Administration Fundamentals"],
        "Pennie": ["Pennie"],
        "Project Management Institute": ["Project Management Ready"],
        "Python Institute": ["PCAP-31-03: Certified Associate in Python Programmer"],
        "The Linux Foundation": ["Linux Foundation Certified IT Associate"],
        "Toon Boom Certified Associate": ["Harmony Advanced", "Harmony Essentials", "Harmony Premium", "Storyboard Pro"],
        "TPQI": ["Level 1", "Level 2"],
        "Unity": ["Unity", "Unity Certified Associate: Artist", "Unity Certified Associate: Programmer", "Unity Certified Expert: Programmer", "Unity Certified Professional: Artist", "Unity Certified Professional: Programmer", "Unity Certified User: Digital Artist", "Unity Certified User: Programmer", "Unity VR Developer"]
    }
    ProductInput['values'] = ProductList_values[selected]


def DisplayStartWindow():
    global driver, check, PasswordInput, Password_var, UsernameInput, Username_var, Category_var, Product_var, ExcelName_var, root, style, ThemeDD, clicked, Background, drop, ProductInput, CategoryInput
    global ProductList
    root = Tk()
    style = Style(theme='forestranger')
    Background = style.colors.bg
    ErrorMessage = ""
    Username_var = StringVar()
    Password_var = StringVar()
    Category_var = StringVar()
    Product_var = StringVar()
    ExcelName_var = StringVar()
    root.title("Authoring Automation Tool")
    driver = None
    root.geometry('1000x800')
    root.configure(background=Background)
    LabelFontStyle = tkFont.Font(family="Helvetica", size=12, weight="bold")
    HeaderFontStyle = tkFont.Font(
        family="Veranda bold", size=20, weight="bold")

    Themes = ["flatly", "cosmo", "litera", "lumen", "pulse", "sandstone", "united", "yeti",
              "superhero", "solar", "cyborg", "darkly", "vaporwave", "carnage", "forestranger", "samhain"]
    CategoriesList = ["Adobe 2021", "Adobe CC", "Adobe CC 2018 (LITA)", "Adobe CS5", "Adobe CS6", "AHIT", "AppInventor", "Apple", "Appraisal",
                      "ASE Entry-Level Certification Program", "ASE Professional Certification", "Autodesk Certified Professional", "Autodesk Certified User", "AWS Certified", "Black Knight", "Coding in Minecraft", "Communication Skills for Business (CSB)", "CompTIA", "Construction", "Digital Skills Programme", "EC-Council",
                      "Entrepreneurship & Small Business", "Exam Prep", "EXIN", "GISP", "Global Digital Literacy", "GMetrix Competency", "IC PHP Developer Fundamentals", "IC3 GS4", "IC3 GS5", "IC3 GS6", "ICC Certifications", "In Development", "Information Technology Specialist", "Introduction to Programming", "Intuit", "Microsoft Certified Educator (MCE)",
                      "Microsoft Certified Fundamentals", "Microsoft Office 2010", "Microsoft Office 2013", "Microsoft Office 2016", "Microsoft Office 2019", "Microsoft Office Corporate Competency", "MTA", "Pennie", "Project Management Institute", "Python Institute", "The Linux Foundation", "Toon Boom Certified Associate", "TPQI", "Unity"]

    # Creating menuFrame
    menuFrame = ttk.Frame(root, style='primary.TFrame')
    menuFrame.pack(pady=50)

    backgroundImage = PhotoImage(file="MicrosoftTeams-image.png")
    backgroundLabel = Label(menuFrame, image=backgroundImage)
    #backgroundLabel.place(x=0, y=0, relwidth=1, relheight=1)

    entryFrame = ttk.Frame(menuFrame, style='secondary.TFrame')
    entryFrame.pack(padx=60, pady=60)

    productFrame = ttk.Frame(menuFrame, style='secondary.TFrame')
    productFrame.pack(padx=10, pady=10)

    ProjectLabel = ttk.Label(menuFrame, text='Question Automation Tool', font=HeaderFontStyle,
                             style='primary.Inverse.TLabel', borderwidth=0)
    ProjectLabel.place(x=120, y=10)

    clicked = StringVar()
    clicked.set("Themes")

    drop = OptionMenu(root, clicked, *Themes,
                      command=(changeTheme)).place(x=850, y=10)

    UsernameLabel = ttk.Label(entryFrame, text='Username :',
                              style='secondary.Inverse.TLabel', font=LabelFontStyle)
    UsernameLabel.grid(row=1, column=1, padx=5, pady=5)

    UsernameInput = ttk.Entry(entryFrame, textvariable=Username_var,
                              style='primary.TEntry', width=25).grid(row=1, column=2, padx=5)

    PasswordLabel = ttk.Label(entryFrame, text='Password :', font=LabelFontStyle,
                              style='secondary.Inverse.TLabel').grid(row=2, column=1, padx=10)

    PasswordInput = ttk.Entry(entryFrame, textvariable=Password_var,
                              style='secondary.TEntry', width=25, show='*')
    PasswordInput.grid(row=2, column=2, padx=20)

    check = ttk.Checkbutton(entryFrame, text='Show Password',
                            command=show, style='success.Roundtoggle.Toolbutton')
    check.grid(row=2, column=3, padx=10, pady=5)

    CategoryLabel = ttk.Label(productFrame, text='Enter Category:',
                              style='secondary.Inverse.TLabel', font=LabelFontStyle).grid(row=1, column=1, padx=10, pady=10)
    CategoryInput = ttk.Combobox(productFrame, textvariable=Category_var,
                                 style='primary.TCombobox', width=30)
    CategoryInput['values'] = CategoriesList
    CategoryInput.grid(row=1, column=2, padx=10, pady=10)
    CategoryInput.bind('<<ComboboxSelected>>', UpdateProductList)

    ProductLabel = ttk.Label(productFrame, text='Enter Product:',
                             style='secondary.Inverse.TLabel', font=LabelFontStyle).grid(row=2, column=1, padx=10, pady=10)
    ProductInput = ttk.Combobox(
        productFrame, textvariable=Product_var, style='primary.TCombobox', width=30)
    ProductInput.grid(row=2, column=2, padx=10, pady=10)

    BrowseButton = ttk.Button(productFrame, text="Browse Files", command=(
        browseFiles), style='danger.TButton').grid(row=3, column=2, padx=5, pady=5)

    Start = ttk.Button(root, text="Start", command=(
        LoginAndOpenQuestionInput), style="info.TButton", width=25).place(x=400, y=400)

    root.mainloop()


def changeTheme(self):
    global style, clicked, root
    themeSet = clicked.get()
    style.theme_use(themeSet)
    Background = style.colors.bg
    root.configure(background=Background)
    root = style.master


def search_for_file_path():
    global tempdir
    currdir = os.getcwd()
    tempdir = filedialog.askopenfilename(
        parent=root, initialdir=currdir, title='Please select a File')
    if len(tempdir) > 0:
        print("You chose: %s" % tempdir)
    return tempdir


def browseFiles():
    search_for_file_path()
    global ExcelFileName
    ExcelFileName = tempdir
    print("\nfile_path_variable = ", ExcelFileName)
    LoadingExcelInfo()


def show():
    PasswordInput.configure(show='')
    check.configure(command=hide, text='Hide Password')


def hide():
    PasswordInput.configure(show='*')
    check.configure(command=show, text='Show Password')


def ErrorWindowDefault():
    messagebox.showerror(title="Error", message=ErrorMessage +", Click OK to cycle to next question or Rerun Sheet")


def ResetWindow():
    root.destroy()
    DisplayStartWindow()


DisplayStartWindow()

# .split('\n')
"""        categorySelected = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="Category-list"]//li[text() = "'+categoryName+'"]'
                                        ))"""


# ~~~~~~~~~~~Made By Diana Cervantes ~~~~~~~~~~~~~ Project: Question Creation Automation
